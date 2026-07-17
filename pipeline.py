"""Pipeline for generating report for ACE-D Aim 3 randomization.

Reads two input CSVs, cleans/reformats each, and writes them to a single
Excel workbook with one tab per CSV.
"""

import argparse
import csv
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.data_source import AxDataSource, NumData, NumVal, StrData, StrRef, StrVal
from openpyxl.chart.legend import LegendEntry
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties, RichTextProperties


def load_csv(path: Path) -> pd.DataFrame:
    """Load a CSV that may have ragged rows (stray trailing commas, or
    rows shorter/longer than the header) without pandas' strict rectangular
    parser rejecting it.
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = [row for row in csv.reader(f) if any(cell.strip() for cell in row)]

    header = [cell.strip() for cell in rows[0]]
    width = max(len(header), max((len(row) for row in rows[1:]), default=0))
    header += [f"col_{i}" for i in range(len(header), width)]

    data = [row + [None] * (width - len(row)) for row in rows[1:]]
    df = pd.DataFrame(data, columns=header)

    for col in df.columns:
        df[col] = df[col].map(lambda v: v.strip() if isinstance(v, str) else v)
        df[col] = df[col].replace({"": None, "NA": None})
        numeric = pd.to_numeric(df[col], errors="coerce")
        if numeric.notna().equals(df[col].notna()):
            df[col] = numeric

    return df


WN_SHEET_NAME = "WebNeuro"
EC_SHEET_NAME = "EtCere"

# Identifying/demographic columns, in the order they should appear first.
WN_IDENTIFYING_COLUMNS = ["ID", "Session", "Age", "Gender", "TestDate"]

# WebNeuro test variables, grouped by test and ordered by the WebNeuro test
# battery's administration order. Each group's color (used to shade that
# test's columns on the norm-score chart) is chosen so it's never adjacent,
# in this sequence, to a similar-looking neighbor.
WN_TEST_GROUPS = [
    ("Motor Tapping", ["tdomnk", "tdomsdk"], "0E76E4"),
    ("Choice Reaction Time", ["chlrrtav"], "8A4100"),
    ("Verbal Memory", ["ctmrec1", "ctmrec2", "ctmrec3", "ctmsco13"], "009C61"),
    ("Emotion Identification", [
        "getcpA", "getcpD", "getcpF", "getcpH", "getcpN", "getcpS",
        "getcrtA", "getcrtD", "getcrtF", "getcrtH", "getcrtN", "getcrtS",
        "gettrtA", "gettrtD", "gettrtF", "gettrtH", "gettrtN", "gettrtS",
    ], "C900E9"),
    ("Digit Span (Forward)", ["digitot", "digitsp"], "BE6900"),
    ("Stroop Word", ["vcrtne", "vi_sco1"], "0095AC"),
    ("Stroop Color", ["vi_difrt", "vcrtne2", "vi_sco2"], "A32638"),
    ("Switching of Attention 1", ["esoadur1", "esoaerr1", "scavr0t1"], "008600"),
    ("Switching of Attention 2", ["esoadur2", "esoaerr2", "scavr0t2"], "B8860B"),
    ("GoNo-Go", ["g2avrtk", "g2errk", "g2fnk", "g2fpk", "g2sdrtk"], "4B32B4"),
    ("Delayed Memory", ["ctmrec4"], "E62B35"),
    ("Emotion Priming", [
        "dgtcnA", "dgtcnD", "dgtcnF", "dgtcnH", "dgtcnS",
        "dgtcrtA", "dgtcrtD", "dgtcrtF", "dgtcrtH", "dgtcrtN", "dgtcrtS",
    ], "23690D"),
    ("N-Back Continuous Performance Test", ["wmacck", "wmfnk", "wmfpk", "wmrtk"], "C84E81"),
    ("Maze", ["emzcompk", "emzerrk", "emzinitk", "emzoverk", "emztrlsk"], "DC4400"),
]

WN_RAW_VARIABLE_ORDER = [name for _, names, _ in WN_TEST_GROUPS for name in names]

# Normed counterparts follow the same order as their raw variables.
WN_NORMED_VARIABLE_ORDER = [f"{name}_norm" for name in WN_RAW_VARIABLE_ORDER]


def order_wn_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Reorder WebNeuro columns: identifying/demographic info first, then
    raw test variables, then normed test variables, each ordered by the
    WebNeuro test battery's administration order (testorder).
    """
    known = set(WN_RAW_VARIABLE_ORDER) | set(WN_NORMED_VARIABLE_ORDER)
    identifying = [c for c in WN_IDENTIFYING_COLUMNS if c in df.columns]
    identifying += [c for c in df.columns if c not in known and c not in identifying]

    ordered = (
        identifying
        + [c for c in WN_RAW_VARIABLE_ORDER if c in df.columns]
        + [c for c in WN_NORMED_VARIABLE_ORDER if c in df.columns]
    )
    return df[ordered]


# Composite score columns appended to the end of the WebNeuro tab: each is
# the row-wise mean of a subset of that test's normed (_norm) variables.
WN_COMPOSITE_GROUPS = [
    ("maze_composite", ["emzcompk", "emzinitk", "emzerrk", "emzoverk", "emztrlsk"]),
    ("gng_composite", ["g2avrtk", "g2errk", "g2fnk", "g2fpk"]),
    ("stroopw_composite", ["vcrtne", "vi_sco1"]),
    ("stroopc_composite", ["vcrtne2", "vi_sco2"]),
    ("swoa_composite", ["esoadur2", "scavr0t2", "esoaerr2"]),
    ("digit_composite", ["digitot", "digitsp"]),
]


def add_wn_composite_scores(df: pd.DataFrame) -> pd.DataFrame:
    """Append one composite-score column per test to the end of the
    dataframe: the row-wise mean of that test's normed variables.
    """
    df = df.copy()
    for name, raw_names in WN_COMPOSITE_GROUPS:
        norm_cols = [f"{n}_norm" for n in raw_names if f"{n}_norm" in df.columns]
        df[name] = df[norm_cols].mean(axis=1) if norm_cols else pd.NA
    return df


# Row-label prefixes identifying the QC metric rows to keep from the EtCere
# export (Signal-to-Noise Ratio and Critical Motion Control rows).
EC_QC_LABEL_PREFIXES = ("Signal-to-Noise Ratio", "Critical Motion Control")

# 1-indexed row (excluding header) of the EtCere export holding the
# "Referenced <participant>" row; its label includes the participant's name
# (auto-populated by the export), so it's kept by fixed row position rather
# than a label match.
EC_REFERENCED_ROW = 10


def filter_ec_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only the "Referenced <participant>" row and the QC metric rows
    from the EtCere export; the rest of its fixed report template
    (raw/global scores, healthy-norm stats, etc.) isn't needed here.
    """
    label_col = df.columns[0]
    is_qc = df[label_col].str.startswith(EC_QC_LABEL_PREFIXES, na=False)
    referenced_row = df.iloc[[EC_REFERENCED_ROW - 1]]
    return pd.concat([referenced_row, df[is_qc]], ignore_index=True)


def write_excel(sheets: dict[str, pd.DataFrame], output_path: Path) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


# One line color for every session; sessions are told apart by dash style
# instead (older session first).
WN_LINE_COLOR = "000000"
WN_SESSION_DASH_STYLES = ["solid", "dash"]
WN_SESSION_LABELS = ["Screening", "Baseline"]

CHART_DATA_SHEET_NAME = "_NormChartData"
CHART_SHEET_NAME = "Norm Score Chart"


def _lighten(hex_color: str, amount: float) -> str:
    """Blend a hex color toward white by `amount` (0 = original, 1 = white),
    for use as a background shade behind foreground lines/text.
    """
    r, g, b = (int(hex_color[i : i + 2], 16) for i in (0, 2, 4))
    return "".join(f"{round(c + (255 - c) * amount):02X}" for c in (r, g, b))


# Background band color for each normed variable's column, keyed by that
# variable's raw (non-"_norm") name, reusing each test's WN_TEST_GROUPS color
# but lightened so foreground lines/labels stay legible on top.
WN_BAND_COLOR_BY_RAW_VARIABLE = {
    name: _lighten(color, 0.75) for _, names, color in WN_TEST_GROUPS for name in names
}


def _cache_series(series, cats_formula: str, categories: list[str], title: str, values: list) -> None:
    """Populate a series's cached title/category/value data.

    openpyxl only ever writes live references (formulas), never the cached
    values Excel normally stores alongside them. Excel appears to treat a
    chart's stored customizations (here, legend-entry deletions — see
    `band_chart.legend.legendEntry` below) as stale and silently drops them
    whenever it has to resolve a series' data from scratch on open, which is
    exactly what happens when that data has no cache. Writing the same cache
    Excel itself would have written keeps those customizations intact.
    """
    series.tx.strRef.strCache = StrData(ptCount=1, pt=[StrVal(idx=0, v=title)])
    series.cat = AxDataSource(
        strRef=StrRef(
            f=cats_formula,
            strCache=StrData(ptCount=len(categories), pt=[StrVal(idx=i, v=v) for i, v in enumerate(categories)]),
        )
    )
    series.val.numRef.numCache = NumData(
        formatCode="General",
        ptCount=len(values),
        pt=[NumVal(idx=i, v=v) for i, v in enumerate(values) if pd.notna(v)],
    )


def add_wn_norm_score_chart(output_path: Path, wn_df: pd.DataFrame) -> None:
    """Add a line chart of the WebNeuro normed scores to the workbook: one
    dashed line per session, plotted against every normed test variable,
    with each variable's column shaded by the test it belongs to.
    """
    categories = [c for c in WN_NORMED_VARIABLE_ORDER if c in wn_df.columns]
    if not categories:
        return
    n_cols = len(categories)

    wb = load_workbook(output_path)

    data_ws = wb.create_sheet(CHART_DATA_SHEET_NAME)
    data_ws.sheet_state = "hidden"
    data_ws.append([None, *categories])
    session_titles = []
    session_values = []
    for row_idx in range(len(wn_df)):
        title = WN_SESSION_LABELS[row_idx] if row_idx < len(WN_SESSION_LABELS) else f"Session {row_idx + 1}"
        values = [wn_df.iloc[row_idx][c] for c in categories]
        data_ws.append([title, *values])
        session_titles.append(title)
        session_values.append(values)
    n_lines = len(wn_df)

    # Background bands need to span the value axis's full range, so pin that
    # range explicitly instead of leaving it to autoscale, and size each
    # band's bar to reach from 0 out to that range's edge.
    data_min = wn_df[categories].min().min()
    data_max = wn_df[categories].max().max()
    bg_row_refs = []
    band_specs = []
    if pd.notna(data_min) and pd.notna(data_max):
        padding = max((data_max - data_min) * 0.1, 0.5)
        y_min, y_max = data_min - padding, data_max + padding
        if y_max > 0:
            data_ws.append(["_bg_pos", *([y_max] * n_cols)])
            bg_row_refs.append(data_ws.max_row)
            band_specs.append(("_bg_pos", [y_max] * n_cols))
        if y_min < 0:
            data_ws.append(["_bg_neg", *([y_min] * n_cols)])
            bg_row_refs.append(data_ws.max_row)
            band_specs.append(("_bg_neg", [y_min] * n_cols))

    cats_ref = Reference(data_ws, min_col=2, max_col=n_cols + 1, min_row=1, max_row=1)
    data_ref = Reference(data_ws, min_col=1, max_col=n_cols + 1, min_row=2, max_row=n_lines + 1)

    # The background bands are a bar chart combined with the line chart:
    # BarChart is the base object (its series render first/lowest, per OOXML
    # plot-area ordering) so it sits behind the LineChart's session lines.
    # Axis configuration lives on the base object since the combine discards
    # the merged-in chart's own axis settings.
    band_chart = BarChart()
    band_chart.height = 14
    band_chart.width = 32
    band_chart.type = "col"
    band_chart.grouping = "stacked"
    band_chart.overlap = 100
    band_chart.gapWidth = 0

    band_chart.x_axis.delete = False
    band_chart.x_axis.axPos = "b"
    # Normed scores straddle 0, so the default "autoZero" crossing would draw
    # the category axis (and its labels) wherever y=0 falls in the plot,
    # i.e. mid-chart. Pin it to the value axis's minimum so it stays at the
    # bottom regardless of the data's sign.
    band_chart.x_axis.crosses = "min"
    band_chart.x_axis.txPr = RichText(
        bodyPr=RichTextProperties(rot=-2700000, vert="horz"),
        # r=[] is required: Paragraph defaults to a single empty text run,
        # which OOXML renders as literal (blank) label text, hiding the
        # auto-generated category labels this txPr is meant to style.
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=700)), endParaRPr=CharacterProperties(sz=700), r=[])],
    )

    band_chart.y_axis.delete = False
    # openpyxl defaults tick marks to explicitly-none, so they must be set
    # explicitly to get visible ticks and labels.
    band_chart.y_axis.majorTickMark = "out"
    band_chart.y_axis.tickLblPos = "nextTo"

    cats_formula = str(cats_ref)

    if bg_row_refs:
        band_chart.y_axis.scaling.min = y_min
        band_chart.y_axis.scaling.max = y_max

        band_ref = Reference(data_ws, min_col=1, max_col=n_cols + 1, min_row=min(bg_row_refs), max_row=max(bg_row_refs))
        band_chart.add_data(band_ref, titles_from_data=True, from_rows=True)
        band_chart.set_categories(cats_ref)
        band_colors = [WN_BAND_COLOR_BY_RAW_VARIABLE[c.removesuffix("_norm")] for c in categories]
        for series, (title, values) in zip(band_chart.series, band_specs):
            series.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            series.dPt = [
                DataPoint(idx=i, spPr=GraphicalProperties(solidFill=color, ln=LineProperties(noFill=True)))
                for i, color in enumerate(band_colors)
            ]
            _cache_series(series, cats_formula, categories, title, values)

    line_chart = LineChart()
    line_chart.add_data(data_ref, titles_from_data=True, from_rows=True)
    line_chart.set_categories(cats_ref)
    for series, dash_style, title, values in zip(line_chart.series, WN_SESSION_DASH_STYLES, session_titles, session_values):
        series.graphicalProperties = GraphicalProperties(
            ln=LineProperties(solidFill=WN_LINE_COLOR, w=19050, prstDash=dash_style)
        )
        series.marker = Marker(symbol="none")
        series.smooth = False
        _cache_series(series, cats_formula, categories, title, values)

    band_chart += line_chart

    band_chart.legend.position = "b"
    # overlay=False reserves dedicated space for the legend below the plot
    # area, so it doesn't sit on top of the rotated x-axis labels.
    band_chart.legend.overlay = False
    # Background band series come first (idx 0, 1, ...); hide them so the
    # legend only lists the session lines.
    band_chart.legend.legendEntry = [LegendEntry(idx=i, delete=True) for i in range(len(bg_row_refs))]

    chart_ws = wb.create_sheet(CHART_SHEET_NAME)
    chart_ws.add_chart(band_chart, "A1")

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("wn_csv", type=Path, help="Path to WebNeuro input CSV")
    parser.add_argument("ec_csv", type=Path, help="Path to EtCere input CSV")
    parser.add_argument("output", type=Path, help="Path to output .xlsx file")
    args = parser.parse_args()

    df1 = order_wn_columns(load_csv(args.wn_csv)).head(2)
    df1 = add_wn_composite_scores(df1)
    df2 = filter_ec_rows(load_csv(args.ec_csv))

    write_excel({WN_SHEET_NAME: df1, EC_SHEET_NAME: df2}, args.output)
    add_wn_norm_score_chart(args.output, df1)
    print(f"Wrote {args.output} with tabs: {WN_SHEET_NAME!r}, {EC_SHEET_NAME!r}")


if __name__ == "__main__":
    main()
